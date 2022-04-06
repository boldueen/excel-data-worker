import openpyxl

# ID
# Статус
# Компания
# ФИО
# Телефон
# Класс
# Отправление
# Назначение
# День недели
# Дата создания
# Время создания
# Дата отложенного заказа
# Время отложенного заказа
# Дата прибытия автомобиля
# Время прибытия автомобиля
# Дата начала следования
# Время начала следования
# Дата завершения
# Время завершения
# Платное ожидание
# ФИО водителя
# Телефон водителя
# Машина
# Рег. номер
# Расстояние
# Время выполнения
# Стоимость
# НДС
# Стоимость с НДС
# Код сотрудника
# Код филиала
# Поставщик
# Стоимость поставщика
# Яндекс: повышенный спрос
# Прибыль (руб)
# Маржа (%)


columns = {
    "ID": 1,
    "status": 2,
    "company": 3,
    "profit": 35,
    "agregator_price": 33,
    "price_nds": 29

}


def is_problem_id(agregator_price, profit):
    if agregator_price == 0 and profit > 0:
        return True
    else:
        return False


def make_report(file):
    wb = openpyxl.load_workbook(filename=file)
    sheet = wb.active

    for row in range(2, sheet.max_row):
        # print(row, sheet.cell(row=row, column=columns.get("price_nds")).value)
        agregator_price = sheet.cell(row=row, column=columns.get("agregator_price")).value
        profit = sheet.cell(row=row, column=columns.get("profit")).value
        status = sheet.cell(row=row, column=columns.get("status")).value

        if is_problem_id(profit=profit, agregator_price=agregator_price):
            print(row, sheet.cell(row=row, column=columns.get("ID")).value)
